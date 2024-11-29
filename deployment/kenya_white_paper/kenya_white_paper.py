# -*- coding: utf-8 -*-
"""
Created on Wed Nov  6 13:53:36 2024

@author: JulianReul
"""

#%%
import time
import numpy as np
import profin as pp
import pandas as pd
from matplotlib import pyplot as plt
from openpyxl import load_workbook

#%%
    "GEOTHERMAL",
    "H2GLOBAL",
    "LOW_EQUITY",
    "COST_OVERRUN",
    "LOW_OFFTAKE",
    "HIGH_PRICE",
    "LOW_CO2_PRICE",
    "LOW_EQUITY",
    "HIGH_INTEREST",
    "NO_COUNTRY_RISK",
    "LOW_BETA",
    "HIGH_INFLATION"


#%% INPUT PARAMETERS

#Initialize storage dicts
STORE_RESULTS = True
FILE_PATH_DEPLOYMENT = "C:\\Users\\JulianReul.AzureAD\\OneDrive - H2Global\\Desktop\\H2Global\\1_Model\\profin_package\\deployment\\kenya_white_paper\\"

SCENARIO_TYPE_LIST = [
    "DEFAULT",
    ]  # 
SENSITIVITY = ["Turkana_South_500_MW", "Turkana_South_100_MW", "Turkana_South_10_MW"]

#Additional parameters
OFFTAKE_PERIOD = 10 #years
ELECTRICITY_SALES = False

#DEFINE DIFFERENT SCENARIOS IN EXCEL SHEET.
data_input = pd.ExcelFile(FILE_PATH_DEPLOYMENT + "kenya_white_paper_scenarios.xlsx")
sheet_names = data_input.sheet_names

for SCENARIO_TYPE in SCENARIO_TYPE_LIST:
    DICT_RESULTS = {}

    for s, sn in enumerate(sheet_names):

        if SCENARIO_TYPE == "DEFAULT" and sn != "Turkana_South_500_MW":
            continue

        #Routine for sensitivity scenarios
        if SCENARIO_TYPE != "DEFAULT":
            #Skip if scenario is not in Turkana South Region
            if sn not in SENSITIVITY:
                continue
            #Rename if sensitivity type is "GEOTHERMAL"
            if SCENARIO_TYPE == "GEOTHERMAL":
                #skip if installed capacity is not 100 MW
                if sn in ["Turkana_South_500_MW", "Turkana_South_10_MW"]:
                    continue
                #rename to link geothermal scenario.
                else:
                    sn = "Turkana_South_100_MW_Grid" 
            
        
        print("SCENARIO: ", sn)
        print("TYPE: ", SCENARIO_TYPE)
        SCENARIO_NAME = SCENARIO_TYPE + "_" + sn
    
        start_time = time.time()
        
        #GENERAL PROJECT DATA
        #____Start of operations
        START_YEAR = 2030
        #____Depreciation period
        DEPRECIATION_PERIOD = 25
        #____Averaged technical lifetime of plant components
        TECHNICAL_LIFETIME = 25
        #____INFLATION
        INFLATION_GLOBAL = 0.03
        INFLATION_DOMESTIC = 0.05
        #____WACC: 13.51%
        #NH3 LHV
        LHV_Ammonia = 5.2 #kWh/kg
        
        #CAPEX and OPEX
        #____Total investment costs (provided as array)
        if SCENARIO_TYPE == "COST_OVERRUN":
            K_INVEST_BASE = pd.read_excel(FILE_PATH_DEPLOYMENT + "kenya_white_paper_scenarios.xlsx", sheet_name=sn, usecols="P").values.flatten()
            K_INVEST = K_INVEST_BASE*1.2
        else:
            K_INVEST = pd.read_excel(FILE_PATH_DEPLOYMENT + "kenya_white_paper_scenarios.xlsx", sheet_name=sn, usecols="P").values.flatten()
    
        #____Terminal value at the end of life
        TERMINAL_VALUE = pd.read_excel(FILE_PATH_DEPLOYMENT + "kenya_white_paper_scenarios.xlsx", sheet_name=sn, usecols="W", skiprows=13, nrows=1).iat[0, 0]
        #____Annual operational costs (provided as array)
        OPEX = pd.read_excel(FILE_PATH_DEPLOYMENT + "kenya_white_paper_scenarios.xlsx", sheet_name=sn, usecols="Q").values.flatten()
        
        
        #ENERGY INPUT / OUTPUT
        #____Energy consumption
        E_in = pd.read_excel(FILE_PATH_DEPLOYMENT + "kenya_white_paper_scenarios.xlsx", sheet_name=sn, usecols="H", skiprows=1, nrows=1).iat[0, 0]*1e+3 #kWh / year
        #____Energy costs
        K_E_in = 0.065 #€/kWh
        
        #____Energy output - Electric surplus 
        if ELECTRICITY_SALES:
            E_out_electric = pd.read_excel(FILE_PATH_DEPLOYMENT + "kenya_white_paper_scenarios.xlsx", sheet_name=sn, usecols="L", skiprows=2, nrows=1).iat[0, 0]*1e+3 #kWh / year
        else:
            E_out_electric = 0
        #____Sales price
        K_E_out_electric_REAL = 0.065 #€/kWh - NOMINAL PRICE
        K_E_out_electric = np.array([K_E_out_electric_REAL*(1+INFLATION_DOMESTIC)**t for t in range(DEPRECIATION_PERIOD)])
        #____Energy output - Electric surplus 
        if SCENARIO_TYPE == "LOW_OFFTAKE":
            E_out_product_BASE = pd.read_excel(FILE_PATH_DEPLOYMENT + "kenya_white_paper_scenarios.xlsx", sheet_name=sn, usecols="L", skiprows=1, nrows=1).iat[0, 0]*1e+3 #kWh / year
            E_out_product = E_out_product_BASE*0.8
        else:
            E_out_product = pd.read_excel(FILE_PATH_DEPLOYMENT + "kenya_white_paper_scenarios.xlsx", sheet_name=sn, usecols="L", skiprows=1, nrows=1).iat[0, 0]*1e+3 #kWh / year
        
        if SCENARIO_TYPE == "HIGH_PRICE":   
            #WtP of local market = x2 of global market.
            CFR_Northwest_Europe = 0.57952*2 #€/kg ammonia in 09/24. Source: S&P Global Commodity Insights
        else:
            CFR_Northwest_Europe = 0.57952 #€/kg ammonia in 09/24. Source: S&P Global Commodity Insights
        if SCENARIO_TYPE == "LOW_CO2_PRICE":
            CO2_PRICE_2030 = 0.1 #€/kg-CO2-eq. in 2023-EUROS. 
            CO2_PRICE_2050 = 0.1 #€/kg-CO2-eq. in 2023-EUROS.
        else:
            CO2_PRICE_2030 = 0.126 #€/kg-CO2-eq. in 2023-EUROS. EY (2023): "CO2-Preisentwicklung im EU ETS – was Unternehmen jetzt beachten müssen", PIK (2024): "The Emerging Endgame: The EU ETS on the Road Towards Climate Neutrality"
            CO2_PRICE_2050 = 0.4 #€/kg-CO2-eq. in 2023-EUROS. EY (2023): "CO2-Preisentwicklung im EU ETS – was Unternehmen jetzt beachten müssen", PIK (2024): "The Emerging Endgame: The EU ETS on the Road Towards Climate Neutrality"
        CO2_INTENSITY_NH3_GRAY = 2.351*0.7 #kg-CO2/kg-NH3 (70% of emissions fall under CBAM)
        Transport_Kenya_Europe = 0.16 #€/kg - Transport & Logistics costs from H2Global tender.
    
        K_E_out_product_REAL_2030 = (CFR_Northwest_Europe - Transport_Kenya_Europe + CO2_INTENSITY_NH3_GRAY*CO2_PRICE_2030) / LHV_Ammonia
        K_E_out_product_REAL_2050 = (CFR_Northwest_Europe - Transport_Kenya_Europe + CO2_INTENSITY_NH3_GRAY*CO2_PRICE_2050) / LHV_Ammonia
        K_E_out_product_REAL = np.linspace(K_E_out_product_REAL_2030, K_E_out_product_REAL_2050, DEPRECIATION_PERIOD)
    
        if SCENARIO_TYPE == "H2GLOBAL":
            H2GLOBAL_OFFTAKE_PRICE_2030 = 0.811*(1+INFLATION_GLOBAL)**(2030-2024) / LHV_Ammonia #€/ton
            K_E_out_product_I = np.array([H2GLOBAL_OFFTAKE_PRICE_2030 for t in range(OFFTAKE_PERIOD)]) #Nominal price does not account for inflation.
            K_E_out_product_II = np.array([K_E_out_product_REAL[t]*(1+INFLATION_GLOBAL)**t for t in range(OFFTAKE_PERIOD, DEPRECIATION_PERIOD)])
            K_E_out_product = np.concatenate((K_E_out_product_I, K_E_out_product_II))
        else:
            #____Sales price
            K_E_out_product = np.array([K_E_out_product_REAL[t]*(1+INFLATION_GLOBAL)**t for t in range(DEPRECIATION_PERIOD)])
        
        print("MIN NH3 [€/ton]:", K_E_out_product.min()*LHV_Ammonia*1000)
        print("MAX NH3 [€/ton]:", K_E_out_product.max()*LHV_Ammonia*1000)
        
        #____Joint quantity
        E_out = E_out_electric + E_out_product
        #____Joint price
        K_E_out = (E_out_electric/E_out)*K_E_out_electric + (E_out_product/E_out)*K_E_out_product
        
        #FINANCIAL METRICS
        #____Equity share
        if SCENARIO_TYPE == "LOW_EQUITY":
            EQUITY_SHARE = 0.25
        else:
            EQUITY_SHARE = 0.4
        #____Interest
        if SCENARIO_TYPE == "HIGH_INTEREST":
            INTEREST = 0.08 #Long-term SWAP-rate (=3%) + 2% credit margin
        else:
            INTEREST = 0.05 #Long-term SWAP-rate (=3%) + 2% credit margin
        #____Tax rate
        CORPORATE_TAX_RATE=0.3 #Damodaran 01/2024
        #____Risk premium
        if SCENARIO_TYPE == "NO_COUNTRY_RISK":
            COUNTRY_RISK_PREMIUM = 0 #0.0951 #Damodaran 01/2024
        else:
            COUNTRY_RISK_PREMIUM = 0.0951
        #____Risk free rate (10y US Government bonds) - NOMINAL TERMS
        R_FREE = 0.045
        #____Return of the mature reference market (S&P500)
        ERP_MATURE = 0.065
        #____unlevered beta factor: 1.058. Lin et al. (2024), DOI: 10.1016/j.ijhydene.2023.07.074
        if SCENARIO_TYPE == "LOW_BETA":
            BETA_UNLEVERED = 0.8 #Damodaran (01/2024), unlevered betas: Chemical (Basic) - 0.82; Oil/Gas (Production & Exploration) - 0.82
        else:
            BETA_UNLEVERED = 1.058#1.058 #Damodaran (01/2024), unlevered betas: Chemical (Basic) - 0.82; Oil/Gas (Production & Exploration) - 0.82
    
        if SCENARIO_TYPE == "RISK":
            ENDOGENOUS_PROJECT_RISK = True
            #____Assumed risks: None. All risks are captured by beta. --> EIN SZENARIO // SENSITIVITÄT
            RISK_PARAM = {
                "K_INVEST": {
                    "distribution": "normal",
                    "scale": K_INVEST * 0.1,
                    "limit": {
                        "min": K_INVEST / 100,
                        "max": K_INVEST * 100
                    },
                    "correlation": {
                        "E_out" : 0.5,
                        "K_E_out" : 0.5,
                    }
                },
                "E_out": {
                    "distribution": "normal",
                    "scale": E_out * 0.1,
                    "limit": {
                        "min": E_out / 100,
                        "max": E_out
                    },
                    "correlation": {
                        "K_INVEST" : 0.5,
                        "K_E_out" : 0.5,
                    }
                },
                "K_E_out": {
                    "distribution": "normal",
                    "scale": K_E_out * 0.2,
                    "limit": {
                        "min": K_E_out / 100,
                        "max": K_E_out
                    },
                    "correlation": {
                        "K_INVEST" : 0.5,
                        "E_out" : 0.5,
                    }
                }
            }
            
        else:
            ENDOGENOUS_PROJECT_RISK = False
            RISK_PARAM = {}
            
        # Scenario calculations
        #Initialize Project-object
        p_example = pp.Project(
            E_in=E_in, #Annual energy consumption in kWh
            E_out=E_out, #Annual production in kWh
            K_E_in=K_E_in, #Average energy purchase price
            K_E_out=K_E_out, #Average energy sales price
            K_INVEST=K_INVEST, #Total investment costs (also provided as an array over depreciation period)
            TERMINAL_VALUE=TERMINAL_VALUE, #Terminal value
            DEPRECIATION_PERIOD=DEPRECIATION_PERIOD, #25 years?
            OPEX=OPEX, #Total operational costs (also possible to provide in array-form)
            EQUITY_SHARE=EQUITY_SHARE,
            COUNTRY_RISK_PREMIUM=COUNTRY_RISK_PREMIUM,
            INTEREST=INTEREST,
            CORPORATE_TAX_RATE=CORPORATE_TAX_RATE,
            RISK_PARAM=RISK_PARAM,
            OBSERVE_PAST=0,
            ENDOGENOUS_PROJECT_RISK=ENDOGENOUS_PROJECT_RISK,
            R_FREE=R_FREE,
            ERP_MATURE=ERP_MATURE,
            TECHNICAL_LIFETIME=TECHNICAL_LIFETIME,
            BETA_UNLEVERED=BETA_UNLEVERED
        )
        
        #CALCULATION OF FINANCIAL METRICS
        
        # Calculate WACC for further calculations.
        WACC = p_example.get_WACC()
        print("____WACC:", WACC.mean())
        
        # Calculate net present value (NPV)
        NPV = p_example.get_NPV(WACC)
        print("____mean NPV:", round(NPV.mean() * 1e-6, 2), " Million €")
        
        LCOE = p_example.get_LCOE(WACC)
        print("____LCOE:", LCOE.mean())
        
        try:
            IRR = p_example.get_IRR(INITIAL_VALUE=0.05)
            print("____IRR:", IRR.mean())
        
            # Calculate the value-at-risk (VaR)
            VaR = p_example.get_VaR(IRR)
            print("____VaR:", VaR)
        except:
            print("____IRR and VaR cannot be determined. - Probably too low.")
              
        DICT_RESULTS[SCENARIO_NAME] = {
            "NPV" : NPV.mean(),
            "LCOE" : LCOE.mean(),
            "IRR" : IRR.mean(),
            "WACC" : WACC.mean()
            }
        
        end_time = time.time()
        delta_time = end_time - start_time
        print("______ELAPSED TIME FOR SINGLE CYCLE: ", delta_time, " sec.")
    
    DICT_RESULTS_pd = pd.DataFrame(DICT_RESULTS)
    try:
        # If the file exists, load the workbook
        book = load_workbook(FILE_PATH_DEPLOYMENT + "RESULTS_KENYA.xlsx")
     
        # Remove sheets with the same names if they exist
        if SCENARIO_TYPE in book.sheetnames:
            # Remove the existing sheet
            std = book[SCENARIO_TYPE]
            book.remove(std)
        
        book.save(FILE_PATH_DEPLOYMENT + "RESULTS_KENYA.xlsx")
        
        # Use ExcelWriter in append mode with the loaded workbook
        with pd.ExcelWriter(FILE_PATH_DEPLOYMENT + "RESULTS_KENYA.xlsx", engine='openpyxl', mode='a') as writer:
            DICT_RESULTS_pd.to_excel(writer, sheet_name=SCENARIO_TYPE, index=True)  # Write DataFrame to the specified sheet
                
    except FileNotFoundError:
        with pd.ExcelWriter(FILE_PATH_DEPLOYMENT + "RESULTS_KENYA.xlsx", engine='openpyxl') as writer:
            DICT_RESULTS_pd.to_excel(writer, sheet_name=SCENARIO_TYPE, index=True)
    
#%%       
 
# Visualize development of NPV over project lifetime
    
LIFETIME_TEMP = TECHNICAL_LIFETIME
WACC_TEMP = WACC

years = np.arange(DEPRECIATION_PERIOD+1)

operating_cashflow, operating_cashflow_std, non_operating_cashflow, non_operating_cashflow_std = p_example.get_cashflows(WACC)

#CASHFLOW
OCF = operating_cashflow
OCF_std = operating_cashflow_std
#____discounting cashflows
CF_discounted = OCF.copy()
CF_std_discounted = OCF_std.copy()
for t in range(DEPRECIATION_PERIOD):
    CF_discounted[t] = OCF[t] / (1+WACC_TEMP)**t
    CF_std_discounted[t] = OCF_std[t] / (1+WACC_TEMP)**t

#ANNUAL NPV
NPV = np.zeros(DEPRECIATION_PERIOD+1)
#____initial invest in year "0"
NPV[0] -= K_INVEST[0].mean()
#____positive cashflows to equity
NPV[1:] = CF_discounted
#____cumulate cashflows
NPV_cum = NPV.cumsum()

#PLOTTING
fig, ax = plt.subplots()
    
line_width = 0.8
#____derive first plot (initial invest)
plot_invest = NPV_cum.copy()
plot_invest[1:] = 0
ax.bar(years, plot_invest, color='orange', width=line_width, label="Invest")

# Get the x-axis limits
delta_x = DEPRECIATION_PERIOD+1
ax.set_xlim(-0.5, DEPRECIATION_PERIOD+0.5)
offset_zero = 0.5 / delta_x

#____derive second plot (positive cashflows)
for year_temp in range(1, DEPRECIATION_PERIOD+1):
    x_position = year_temp/delta_x
    ax.axhline(y=NPV_cum[year_temp], 
                color='black', 
                xmin=x_position - (1/delta_x)*0.4 + offset_zero, 
                xmax=x_position + (1/delta_x)*0.4 + offset_zero, 
                linestyle='-')

#____derive third and fourth plot (terminal values)
#Accounting for open principal payments.
REPAYMENT_PERIOD = DEPRECIATION_PERIOD
INVEST_TEMP = K_INVEST[0]
OPEN_PRINCIPAL = 0
terminal_value = TERMINAL_VALUE / (1+WACC_TEMP)**DEPRECIATION_PERIOD
plot_terminal_value = np.zeros(DEPRECIATION_PERIOD+1)
plot_npv = np.zeros(DEPRECIATION_PERIOD+1)

if NPV_cum[-1] < 0:
    if terminal_value+NPV_cum[-1] > 0:
        #plot positive NPV in green
        plot_npv[-1] = terminal_value+NPV_cum[-1]
        ax.bar(years, plot_npv, color='green', width=line_width, label="Positive NPV")
        #plot terminal value add on until NPV=0
        plot_terminal_value[-1] = -NPV_cum[-1]
        ax.bar(years, plot_terminal_value, bottom=NPV_cum[-1], color='grey', width=line_width, label="Terminal value-NPV")
    else:
        #plot negative NPV in red
        plot_npv[-1] = -(terminal_value+NPV_cum[-1])
        ax.bar(years, plot_npv, bottom=NPV_cum[-1]+terminal_value, color='red', width=line_width, label="Negative NPV")
        #plot terminal value add on until NPV=0
        plot_terminal_value[-1] = terminal_value
        ax.bar(years, plot_terminal_value, bottom=NPV_cum[-1], color='grey', width=line_width, label="Terminal value")
else:
    #plot terminal value on top of positive NPV in green 
    plot_npv[-1] = NPV_cum[-1]+terminal_value
    ax.bar(years, plot_npv, bottom=0, color='green', width=line_width, label="Positive NPV")
    
ax.set_xlabel('Years')
ax.set_ylabel('Net present value [US$]')
#ax.ylim(0, 4*1e8)
ax.set_title(s)
ax.legend(loc="lower right")
ax.grid(True)
xtick_position = [1, 6, 11, 16, 21, 26, 31]
xtick_label = [2030, 2035, 2040, 2045, 2050, 2055, 2060]
plt.xticks(xtick_position, xtick_label)
plt.show()
